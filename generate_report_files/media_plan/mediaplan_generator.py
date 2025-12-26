#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Генератор медиаплана - скрипт для формирования Excel-документов по отчетам из БД
"""

import os
import json
import io

import psycopg2
from minio import Minio
from dotenv import load_dotenv
from typing import Dict, List, Optional, Any
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Side, Border

from utils.postprocessing_report_file import write_s3path_to_bd

# Загружаем переменные окружения
load_dotenv()

class MediaPlanGenerator:
    def __init__(self):
        """Инициализация подключений к БД и MinIO"""
        # Настройки БД
        self.db_config = {
            'host': os.getenv('DB_HOST', 'localhost'),
            'port': os.getenv('DB_PORT', '5432'),
            'database': os.getenv('DB_NAME'),
            'user': os.getenv('DB_USER'),
            'password': os.getenv('DB_PASSWORD')
        }
        
        # Настройки MinIO
        self.minio_client = Minio(
            endpoint=os.getenv('S3_ENDPOINT_URL', 'minio.upk-mos.ru'),
            access_key=os.getenv('S3_ACCESS_KEY'),
            secret_key=os.getenv('S3_SECRET_KEY'),
            secure=os.getenv('S3_SECURE', 'False').lower() == 'true'
        )
        
        self.bucket_name = os.getenv('S3_BUCKET_NAME', 'dit-services-dev')
        
        # Папка для результатов
        self.output_folder = 'mediaplan_results'
        self._ensure_output_folder()

        # Цвет заливки для заголовков
        self.header_fill = PatternFill(start_color='FFF2CD', end_color='FFF2CD', fill_type='solid')
        
        # Счетчик для групп ключей и словарь соответствия
        self.key_group_counter = 1
        self.campaign_group_numbers = {}

    def _ensure_output_folder(self):
        """Создать папку для результатов, если её нет"""
        if not os.path.exists(self.output_folder):
            os.makedirs(self.output_folder)
            print(f"📁 Создана папка для результатов: {self.output_folder}")
        else:
            print(f"📁 Папка для результатов: {self.output_folder}")

    def get_project_names(self, project_ids: List[int]) -> Dict[int, str]:
        """Получить названия проектов по их ID из БД"""
        try:
            conn = psycopg2.connect(**self.db_config)
            cursor = conn.cursor()
            
            # Устанавливаем схему по умолчанию
            cursor.execute("SET search_path TO gen_report_context_contracts, public;")
            
            # Получаем названия проектов
            placeholders = ','.join(['%s'] * len(project_ids))
            query = f"""
            SELECT p.id, p.name as project_name
            FROM projects p
            WHERE p.id IN ({placeholders})
            """
            
            cursor.execute(query, project_ids)
            project_names = {row[0]: row[1] for row in cursor.fetchall()}
            
            cursor.close()
            conn.close()
            
            return project_names
            
        except Exception as e:
            print(f"❌ Ошибка при получении названий проектов: {e}")
            return {}

    def get_pending_reports(self) -> List[Dict]:
        """Получить отчеты со статусом 1 (готовые к обработке)"""
        try:
            print(f"🔌 Подключение к БД: {self.db_config['host']}:{self.db_config['port']}/{self.db_config['database']}")
            conn = psycopg2.connect(**self.db_config)
            cursor = conn.cursor()
            
            # Устанавливаем схему по умолчанию
            cursor.execute("SET search_path TO gen_report_context_contracts, public;")
            print("✓ Схема установлена: gen_report_context_contracts")
            
            # Получаем отчеты со статусом 1 и данные заявки
            query = """
            SELECT 
                r.id, 
                r.id_contracts, 
                r.id_requests, 
                c.number_contract, 
                c.subject_contract,
                req.campany_yandex_direct
            FROM reports r
            JOIN contracts c ON r.id_contracts = c.id
            JOIN requests req ON r.id_requests = req.id
            WHERE r.id_status = 1 AND (r.is_deleted = false OR r.is_deleted IS NULL)
            ORDER BY r.create_entry DESC
            """
            
            print(f"🔍 Выполняем запрос...")
            cursor.execute(query)
            reports = []
            
            for row in cursor.fetchall():
                reports.append({
                    'id': row[0],
                    'id_contracts': row[1],
                    'id_requests': row[2],
                    'number_contract': row[3],
                    'subject_contract': row[4],
                    'campaign_ids': row[5]  # JSONB с id кампаний
                })
            
            print(f"✅ Найдено отчетов для обработки: {len(reports)}")
            
            cursor.close()
            conn.close()
            
            return reports
            
        except Exception as e:
            print(f"❌ Ошибка при получении отчетов из БД: {e}")
            return []

    def load_file_from_minio(self, report_id: int, filename: str) -> Optional[Dict]:
        """Загрузить JSON файл из MinIO для конкретного отчета"""
        try:
            folder_path = f"gen_report_context_contracts/data_yandex_direct/{report_id}_результаты"
            object_path = f"{folder_path}/{filename}"
            
            # Проверяем существование объекта
            if self.minio_client.stat_object(self.bucket_name, object_path):
                # Загружаем объект
                response = self.minio_client.get_object(self.bucket_name, object_path)
                content = response.read().decode('utf-8')
                data = json.loads(content)
                print(f"✓ Загружен файл {filename}")
                return data
            else:
                print(f"⚠ Файл {filename} не найден")
                return None
                
        except Exception as e:
            print(f"✗ Ошибка при загрузке {filename}: {e}")
            return None

    def get_unique_ad_combinations(self, campaign_id: int, ads_data: Dict) -> List[Dict]:
        """Получить уникальные комбинации заголовок-текст из объявлений"""
        try:
            if not ads_data or not ads_data.get('result'):
                return []
            
            combinations = set()
            result = []
            for ad in ads_data.get('result', {}).get('Ads', []):
                if ad.get('CampaignId') == campaign_id:
                    text_ad = ad.get('TextAd', {})
                    title = text_ad.get('Title', '')
                    text = text_ad.get('Text', '')
                    if title and text:
                        combination = (title, text)
                        if combination not in combinations:
                            combinations.add(combination)
                            result.append({
                                'title': title,
                                'text': text
                            })
            return result
        except Exception as e:
            print(f"  ⚠ Ошибка при получении комбинаций объявлений: {e}")
            return []

    def get_campaign_callouts(self, campaign_id: int, extensions_data: Dict, ads_data: Dict) -> List[str]:
        """Получить уникальные уточнения для кампании"""
        try:
            if not extensions_data or not isinstance(extensions_data, dict) or not ads_data:
                return []
            
            # Сначала найдем все ID уточнений, используемые в объявлениях кампании
            extension_ids = set()
            for ad in ads_data.get('result', {}).get('Ads', []):
                if ad.get('CampaignId') == campaign_id:
                    ad_extensions = ad.get('TextAd', {}).get('AdExtensions', [])
                    for ext in ad_extensions:
                        if ext.get('Type') == 'CALLOUT':
                            extension_ids.add(ext.get('AdExtensionId'))
            
            # Если нет уточнений, возвращаем пустой список
            if not extension_ids:
                return []
            
            # Теперь найдем тексты уточнений по их ID, сохраняя порядок из файла
            callouts = []
            seen = set()  # Для проверки уникальности
            batch_data = next(iter(extensions_data.values()), {})
            if not batch_data or not batch_data.get('result'):
                return []
            
            for ext in batch_data.get('result', {}).get('AdExtensions', []):
                if ext.get('Id') in extension_ids and ext.get('Type') == 'CALLOUT':
                    callout_text = ext.get('Callout', {}).get('CalloutText', '')
                    if callout_text and callout_text not in seen:
                        seen.add(callout_text)
                        callouts.append(callout_text)
            
            return callouts  # Возвращаем список в порядке появления в файле
        except Exception as e:
            print(f"  ⚠ Ошибка при получении уточнений: {e}")
            return []

    def get_campaign_adgroups(self, campaign_id: int, adgroups_data: Dict) -> List[str]:
        """Получить уникальные названия групп объявлений для кампании"""
        try:
            if not adgroups_data or not adgroups_data.get('result'):
                return []
            
            group_names = set()
            for group in adgroups_data.get('result', {}).get('AdGroups', []):
                if group.get('CampaignId') == campaign_id:
                    name = group.get('Name', '')
                    if name:
                        group_names.add(name)
            return sorted(list(group_names))
        except Exception as e:
            print(f"  ⚠ Ошибка при получении групп объявлений: {e}")
            return []

    def get_campaign_sitelinks(self, campaign_id: int, sitelinks_data: Dict, ads_data: Dict) -> List[Dict]:
        """Получить уникальные быстрые ссылки для кампании"""
        try:
            if not sitelinks_data or not isinstance(sitelinks_data, dict) or not ads_data:
                return []
            
            # Сначала найдем все ID наборов быстрых ссылок, используемые в объявлениях кампании
            sitelink_set_ids = set()
            for ad in ads_data.get('result', {}).get('Ads', []):
                if ad.get('CampaignId') == campaign_id:
                    sitelink_set_id = ad.get('TextAd', {}).get('SitelinkSetId')
                    if sitelink_set_id:
                        sitelink_set_ids.add(str(sitelink_set_id))
            
            # Если нет наборов быстрых ссылок, возвращаем пустой список
            if not sitelink_set_ids:
                return []
            
            sitelinks = []
            seen = set()
            
            # Проходим по всем наборам данных
            for set_id, set_data in sitelinks_data.items():
                # Проверяем, что это нужный нам набор быстрых ссылок
                if set_id not in sitelink_set_ids:
                    continue
                
                if not isinstance(set_data, dict):
                    continue
                
                # Получаем список наборов быстрых ссылок
                sitelinks_sets = set_data.get('result', {}).get('SitelinksSets', [])
                if not sitelinks_sets:
                    continue
                
                # Обрабатываем каждый набор
                for sitelinks_set in sitelinks_sets:
                    if not isinstance(sitelinks_set, dict):
                        continue
                    
                    # Получаем список ссылок
                    links = sitelinks_set.get('Sitelinks', [])
                    if not isinstance(links, list):
                        continue
                    
                    # Обрабатываем каждую ссылку
                    for link in links:
                        if not isinstance(link, dict):
                            continue
                        
                        title = link.get('Title', '')
                        description = link.get('Description', '')
                        href = link.get('Href', '')
                        
                        # Создаем уникальный ключ для комбинации
                        key = (title, description, href)
                        if key not in seen and all([title, description, href]):
                            seen.add(key)
                            sitelinks.append({
                                'title': title,
                                'description': description,
                                'href': href
                            })
            
            return sitelinks
        except Exception as e:
            print(f"  ⚠ Ошибка при получении быстрых ссылок: {e}")
            return []

    def get_unique_hrefs(self, campaign_id: int, ads_data: Dict) -> List[str]:
        """Получить список уникальных ссылок из объявлений кампании"""
        try:
            if not ads_data or not ads_data.get('result'):
                return []
            
            hrefs = set()
            for ad in ads_data.get('result', {}).get('Ads', []):
                if ad.get('CampaignId') == campaign_id:
                    href = ad.get('TextAd', {}).get('Href')
                    if href:
                        hrefs.add(href)
            
            return sorted(list(hrefs))
        except Exception as e:
            print(f"  ⚠ Ошибка при получении ссылок: {e}")
            return []

    def get_negative_keywords(self, campaign: Dict) -> List[str]:
        """Получить список минус-слов для кампании"""
        try:
            if not campaign:
                return []
                
            negative_keywords = campaign.get('NegativeKeywords')
            if not negative_keywords:
                return []
                
            items = negative_keywords.get('Items', [])
            if not items:
                return []
                
            return items  # Возвращаем список как есть, очистку кавычек делаем при добавлении в Excel
        except Exception as e:
            print(f"  ⚠ Ошибка при получении минус-слов: {e}")
            return []

    def get_campaign_keywords(self, campaign_id: int, keywords_data: Dict) -> List[str]:
        """Получить список ключевых фраз для кампании"""
        if not keywords_data or not keywords_data.get('result'):
            return []
        
        keywords = []
        for keyword in keywords_data.get('result', {}).get('Keywords', []):
            if keyword.get('CampaignId') == campaign_id:
                # Убираем кавычки из ключевой фразы
                keyword_text = keyword.get('Keyword', '').strip('"')
                # Пропускаем автотаргетинг
                if keyword_text and '---autotargeting' not in keyword_text:
                    keywords.append(keyword_text)
        
        return keywords  # Возвращаем как есть, без сортировки

    def is_keyword_campaign(self, campaign_id: int, keywords_data: Dict) -> bool:
        """
        Определить, является ли кампания ключевой
        Если есть хотя бы одна ключевая фраза (кроме автотаргетинга), считаем кампанию ключевой
        """
        keywords = self.get_campaign_keywords(campaign_id, keywords_data)
        return len(keywords) > 0

    def categorize_campaigns(self, campaigns_data: Dict, request_campaign_ids: Any, keywords_data: Dict) -> Dict[int, Dict[str, Dict[str, List[Dict]]]]:
        """
        Разделить кампании на категории и проекты:
        - Сначала группируем по project_id
        - Внутри каждого проекта делим на:
          - РСЯ (ключи/интересы)
          - Поиск (ключи/интересы)
          - МК (пока пропускаем)
        """
        if not campaigns_data or not campaigns_data.get('result'):
            print("⚠ Нет данных кампаний")
            return {}
        
        all_campaigns = campaigns_data.get('result', {}).get('Campaigns', [])
        print(f"📊 Всего кампаний в данных: {len(all_campaigns)}")
        
        # Получаем список кампаний с их project_id из заявки
        allowed_campaigns = {}  # Dict[campaign_id, project_id]
        if request_campaign_ids:
            if isinstance(request_campaign_ids, dict) and 'campaigns' in request_campaign_ids:
                for campaign in request_campaign_ids['campaigns']:
                    if isinstance(campaign, dict):
                        campaign_id = campaign.get('id')
                        project_id = campaign.get('project_id')
                        if campaign_id is not None and project_id is not None:
                            allowed_campaigns[int(campaign_id)] = int(project_id)
            elif isinstance(request_campaign_ids, list):
                for item in request_campaign_ids:
                    if isinstance(item, dict):
                        campaign_id = item.get('id')
                        project_id = item.get('project_id')
                        if campaign_id is not None and project_id is not None:
                            allowed_campaigns[int(campaign_id)] = int(project_id)

        print(f"📋 Кампании и их проекты из заявки: {allowed_campaigns}")

        # Группируем кампании по проектам
        projects = {}  # Dict[project_id, categories]
        
        # Фильтруем и категоризируем кампании
        for campaign in all_campaigns:
            campaign_id = campaign.get('Id')
            campaign_name = campaign.get('Name', '')
            
            # Проверяем, что кампания есть в списке из заявки
            if not allowed_campaigns or campaign_id not in allowed_campaigns:
                continue
                
            project_id = allowed_campaigns[campaign_id]
            
            # Инициализируем структуру для проекта, если её ещё нет
            if project_id not in projects:
                projects[project_id] = {
                    'rsy': {
                        'keywords': [],  # РСЯ кампании с ключами
                        'interests': []  # РСЯ кампании с интересами
                    },
                    'search': {
                        'keywords': [],  # Поисковые кампании с ключами
                        'interests': []  # Поисковые кампании с интересами
                    },
                    'mk': []  # МК (пока не обрабатываем)
                }
            
            # Определяем категорию по имени
            if '/РСЯ/' in campaign_name:
                # Определяем тип кампании (ключи/интересы)
                if self.is_keyword_campaign(campaign_id, keywords_data):
                    projects[project_id]['rsy']['keywords'].append(campaign)
                    print(f"  ✓ Проект {project_id}: Найдена РСЯ-кампания с ключами: {campaign_name}")
                else:
                    projects[project_id]['rsy']['interests'].append(campaign)
                    print(f"  ✓ Проект {project_id}: Найдена РСЯ-кампания с интересами: {campaign_name}")
            elif '/Поиск/' in campaign_name:
                # Определяем тип кампании (ключи/интересы)
                if self.is_keyword_campaign(campaign_id, keywords_data):
                    projects[project_id]['search']['keywords'].append(campaign)
                    print(f"  ✓ Проект {project_id}: Найдена поисковая кампания с ключами: {campaign_name}")
                else:
                    projects[project_id]['search']['interests'].append(campaign)
                    print(f"  ✓ Проект {project_id}: Найдена поисковая кампания с интересами: {campaign_name}")
            elif '/МК/' in campaign_name:
                projects[project_id]['mk'].append(campaign)
                print(f"  ✓ Проект {project_id}: Найдена МК-кампания: {campaign_name} (пропускается)")

        print(f"✅ Категоризация по проектам завершена:")
        for project_id, categories in projects.items():
            print(f"\nПроект {project_id}:")
            print(f"  РСЯ (ключи): {len(categories['rsy']['keywords'])} кампаний")
            print(f"  РСЯ (интересы): {len(categories['rsy']['interests'])} кампаний")
            print(f"  Поиск (ключи): {len(categories['search']['keywords'])} кампаний")
            print(f"  Поиск (интересы): {len(categories['search']['interests'])} кампаний")
            print(f"  МК: {len(categories['mk'])} кампаний (пропускаются)")

        return projects

    def adjust_column_widths(self, ws):
        """Настроить ширину столбцов"""
        ws.column_dimensions['A'].width = 56  # Первый столбец шириной 56
        ws.column_dimensions['B'].width = 15  # Минус-слова
        ws.column_dimensions['C'].width = 25  # Посадочная страница

    def sanitize_sheet_name(self, sheet_name: str, max_length: int = 31) -> str:
        """
        Создать безопасное имя листа Excel:
        - Убрать недопустимые символы
        - Обрезать до максимальной длины
        """
        # Заменяем недопустимые символы
        invalid_chars = '[]:*?/\\'
        for char in invalid_chars:
            sheet_name = sheet_name.replace(char, ' ')
        
        # Если имя слишком длинное, просто обрезаем
        if len(sheet_name) > max_length:
            sheet_name = sheet_name[:max_length]
            
        return sheet_name.strip()

    def extract_campaign_type(self, campaign_name: str) -> str:
        """
        Извлечь тип кампании (РСЯ/Поиск)
        """
        if '/РСЯ/' in campaign_name:
            return 'РСЯ'
        elif '/Поиск/' in campaign_name:
            return 'Поиск'
        return ''

    def create_mediaplan_excel(self, report_id: int, projects: Dict[int, Dict[str, Dict[str, List[Dict]]]], keywords_data: Dict, ads_data: Dict, extensions_data: Dict, sitelinks_data: Dict, adgroups_data: Dict, output_path: str) -> bool:
        """Создать Excel-документ с медиапланом"""
        try:
            # Создаем новую книгу Excel
            wb = openpyxl.Workbook()
            
            # Удаляем стандартный лист
            wb.remove(wb.active)
            
            # Сначала присваиваем номера группам для всех кампаний
            self.campaign_group_numbers = {}  # Сбрасываем словарь
            self.key_group_counter = 1  # Сбрасываем счетчик
            
            # Получаем названия проектов
            project_names = self.get_project_names(list(projects.keys()))
            
            # Сортируем проекты по ID для предсказуемого порядка
            sorted_project_ids = sorted(projects.keys())
            
            # Проходим по всем проектам и кампаниям для нумерации
            for project_id in sorted_project_ids:
                categories = projects[project_id]
                for category in ['rsy', 'search']:
                    campaigns = categories[category]['keywords']
                    for campaign in campaigns:
                        campaign_id = campaign.get('Id')
                        if campaign_id not in self.campaign_group_numbers:
                            self.campaign_group_numbers[campaign_id] = self.key_group_counter
                            self.key_group_counter += 1
            
            # Теперь создаем листы Excel для каждого проекта
            for project_id in sorted_project_ids:
                categories = projects[project_id]
                project_name = project_names.get(project_id, f"Проект {project_id}")
                
                for category in ['rsy', 'search']:
                    # Обрабатываем только кампании с ключами
                    campaigns = categories[category]['keywords']
                    if not campaigns:
                        continue
                    
                    category_name = 'РСЯ' if category == 'rsy' else 'Поиск'
                    
                    # Создаем листы для ключей
                    sheet_name = self.sanitize_sheet_name(f"Ключи - {category_name} - {project_name}")
                    ws = wb.create_sheet(sheet_name)
                    
                    # Добавляем заголовки
                    ws['A1'] = "Название / Тип"
                    ws['B1'] = "Минус-слова"
                    ws['C1'] = "Посадочная страница"
                    
                    # Применяем заливку и жирный шрифт к заголовкам
                    for cell in ws[1]:
                        cell.fill = self.header_fill
                        cell.font = Font(bold=True)
                    
                    # Фиксируем первую строку
                    ws.freeze_panes = 'A2'
                    
                    current_row = 2
                    
                    # Обрабатываем каждую кампанию
                    for campaign in campaigns:
                        campaign_id = campaign.get('Id')
                        campaign_name = campaign.get('Name', '')
                        
                        # Добавляем название кампании как есть
                        cell = ws.cell(row=current_row, column=1, value=campaign_name)
                        cell.fill = self.header_fill
                        cell.font = Font(bold=True)
                        
                        # Оставляем пустую ячейку для минус-слов в заголовке кампании без стилей
                        ws.cell(row=current_row, column=2, value="")
                        
                        # Настраиваем ширину столбцов
                        self.adjust_column_widths(ws)
                        
                        # Используем тот же номер группы, что и на листе "Ключи"
                        campaign_id = campaign.get('Id')
                        group_number = self.campaign_group_numbers.get(campaign_id, self.key_group_counter)
                        
                        # Добавляем "Группа ключей N", используя ранее присвоенный номер
                        current_row += 1
                        group_number = self.campaign_group_numbers[campaign_id]
                        cell = ws.cell(row=current_row, column=1, value=f"Группа ключей {group_number}")
                        cell.fill = self.header_fill
                        cell.font = Font(bold=True)
                        
                        # Получаем все данные
                        keywords = self.get_campaign_keywords(campaign_id, keywords_data)
                        minus_words = self.get_negative_keywords(campaign)
                        hrefs = self.get_unique_hrefs(campaign_id, ads_data)
                        
                        # Определяем начальную строку для контента
                        content_start_row = current_row + 1
                        
                        # Добавляем ключевые фразы
                        current_row = content_start_row
                        for keyword in keywords:
                            ws.cell(row=current_row, column=1, value=keyword)
                            current_row += 1
                        
                        # Добавляем минус-слова в столбик
                        current_row = content_start_row
                        for minus_word in minus_words:
                            ws.cell(row=current_row, column=2, value=minus_word.strip('"'))
                            current_row += 1
                        
                        # Добавляем ссылки в столбик
                        current_row = content_start_row
                        for href in hrefs:
                            ws.cell(row=current_row, column=3, value=href)
                            current_row += 1
                        
                        # Обновляем current_row до максимального значения
                        current_row = max(
                            content_start_row + len(keywords),
                            content_start_row + len(minus_words),
                            content_start_row + len(hrefs)
                        )
                    
                    # Создаем лист медиаплана для кампаний с ключами
                    sheet_name = self.sanitize_sheet_name(f"Медиаплан - {category_name} - {project_name}")
                    ws = wb.create_sheet(sheet_name)
                    
                    # Настраиваем ширину столбцов
                    ws.column_dimensions['A'].width = 30  # Заголовок
                    ws.column_dimensions['B'].width = 30  # Текст
                    ws.column_dimensions['C'].width = 30  # Уточнения
                    ws.column_dimensions['D'].width = 30  # Быстрые ссылки
                    ws.column_dimensions['E'].width = 30  # Описание быстрых ссылок
                    ws.column_dimensions['F'].width = 30  # Адреса быстрых ссылок
                    
                    # Создаем лист для кампаний с интересами
                    if categories[category]['interests']:
                        sheet_name = self.sanitize_sheet_name(f"Интересы - {category_name} - {project_name}")
                        ws_interests = wb.create_sheet(sheet_name)
                        
                        # Настраиваем ширину столбцов
                        ws_interests.column_dimensions['A'].width = 30  # Интересы
                        ws_interests.column_dimensions['B'].width = 30  # Посадочная страница
                        ws_interests.column_dimensions['C'].width = 30  # Заголовок
                        ws_interests.column_dimensions['D'].width = 30  # Текст
                        ws_interests.column_dimensions['E'].width = 30  # Уточнения
                        ws_interests.column_dimensions['F'].width = 30  # Быстрые ссылки
                        ws_interests.column_dimensions['G'].width = 30  # Описание быстрых ссылок
                        ws_interests.column_dimensions['H'].width = 30  # Адреса быстрых ссылок
                        
                        current_row = 1
                        
                        # Обрабатываем каждую кампанию с интересами
                        for campaign in categories[category]['interests']:
                            campaign_id = campaign.get('Id')
                            campaign_name = campaign.get('Name', '')
                            
                            # Добавляем название кампании
                            cell = ws_interests.cell(row=current_row, column=1, value=campaign_name)
                            cell.fill = self.header_fill
                            cell.font = Font(bold=True)
                            
                            # Добавляем заголовки столбцов
                            current_row += 1
                            headers = ["Интересы", "Посадочная страница", "Заголовок", "Текст", 
                                     "Уточнения", "Быстрые ссылки", "Описание быстрых ссылок", "Адреса быстрых ссылок"]
                            for col, header in enumerate(headers, 1):
                                cell = ws_interests.cell(row=current_row, column=col, value=header)
                                cell.fill = self.header_fill
                                cell.font = Font(bold=True)
                            
                            # Получаем все данные
                            adgroups = self.get_campaign_adgroups(campaign_id, adgroups_data)
                            hrefs = self.get_unique_hrefs(campaign_id, ads_data)
                            ad_combinations = self.get_unique_ad_combinations(campaign_id, ads_data)
                            callouts = self.get_campaign_callouts(campaign_id, extensions_data, ads_data)
                            sitelinks = self.get_campaign_sitelinks(campaign_id, sitelinks_data, ads_data)
                            
                            # Добавляем данные
                            content_row = current_row + 1
                            max_rows = max(len(adgroups), len(hrefs), len(ad_combinations), len(sitelinks), 1)
                            
                            # Добавляем интересы (названия групп)
                            for i, group_name in enumerate(adgroups):
                                ws_interests.cell(row=content_row + i, column=1, value=group_name)
                            
                            # Добавляем посадочные страницы
                            for i, href in enumerate(hrefs):
                                ws_interests.cell(row=content_row + i, column=2, value=href)
                            
                            # Добавляем комбинации заголовок-текст
                            for i, ad in enumerate(ad_combinations):
                                ws_interests.cell(row=content_row + i, column=3, value=ad['title'])
                                ws_interests.cell(row=content_row + i, column=4, value=ad['text'])
                            
                            # Добавляем уточнения
                            if callouts:
                                # Объединяем ячейки для уточнений
                                if max_rows > 1:
                                    ws_interests.merge_cells(
                                        start_row=content_row,
                                        start_column=5,
                                        end_row=content_row + max_rows - 1,
                                        end_column=5
                                    )
                                cell = ws_interests.cell(row=content_row, column=5, value="\n".join(callouts))
                                cell.alignment = Alignment(wrapText=True, vertical='top')
                            
                            # Добавляем быстрые ссылки
                            for i, link in enumerate(sitelinks):
                                ws_interests.cell(row=content_row + i, column=6, value=link['title'])
                                ws_interests.cell(row=content_row + i, column=7, value=link['description'])
                                ws_interests.cell(row=content_row + i, column=8, value=link['href'])
                            
                            # Обновляем current_row для следующей кампании
                            current_row = content_row + max_rows
                    
                    current_row = 1
                    
                    # Обрабатываем каждую кампанию
                    for campaign in campaigns:
                        campaign_id = campaign.get('Id')
                        campaign_name = campaign.get('Name', '')
                        
                        # Добавляем номер группы ключей, используя ранее присвоенный номер
                        campaign_id = campaign.get('Id')
                        group_number = self.campaign_group_numbers[campaign_id]
                        cell = ws.cell(row=current_row, column=1, value=f"Группа ключей {group_number}")
                        cell.fill = self.header_fill
                        cell.font = Font(bold=True)
                        
                        # Добавляем название кампании
                        current_row += 1
                        cell = ws.cell(row=current_row, column=1, value=campaign_name)
                        cell.fill = self.header_fill
                        cell.font = Font(bold=True)
                        
                        # Добавляем заголовки столбцов
                        current_row += 1
                        headers = ["Заголовок", "Текст", "Уточнения", "Быстрые ссылки", "Описание быстрых ссылок", "Адреса быстрых ссылок"]
                        for col, header in enumerate(headers, 1):
                            cell = ws.cell(row=current_row, column=col, value=header)
                            cell.fill = self.header_fill
                            cell.font = Font(bold=True)
                        
                        # Получаем данные для кампании
                        ad_combinations = self.get_unique_ad_combinations(campaign_id, ads_data)
                        callouts = self.get_campaign_callouts(campaign_id, extensions_data, ads_data)
                        sitelinks = self.get_campaign_sitelinks(campaign_id, sitelinks_data, ads_data)
                        
                        # Добавляем данные
                        content_row = current_row + 1
                        
                        # Добавляем комбинации заголовок-текст
                        for ad in ad_combinations:
                            ws.cell(row=content_row, column=1, value=ad['title'])
                            ws.cell(row=content_row, column=2, value=ad['text'])
                            content_row += 1
                        
                        # Добавляем уточнения и объединяем ячейки
                        if callouts:
                            # Вычисляем количество строк для объединения
                            max_rows = max(len(ad_combinations), len(sitelinks), 1)
                            
                            # Объединяем ячейки
                            if max_rows > 1:
                                ws.merge_cells(
                                    start_row=current_row + 1,
                                    start_column=3,
                                    end_row=current_row + max_rows,
                                    end_column=3
                                )
                            
                            # Добавляем текст с переносами строк
                            cell = ws.cell(row=current_row + 1, column=3, value="\n".join(callouts))
                            # Включаем перенос строк и выравнивание по верху
                            cell.alignment = Alignment(wrapText=True, vertical='top')
                        
                        # Добавляем быстрые ссылки
                        content_row = current_row + 1
                        for link in sitelinks:
                            ws.cell(row=content_row, column=4, value=link['title'])
                            ws.cell(row=content_row, column=5, value=link['description'])
                            ws.cell(row=content_row, column=6, value=link['href'])
                            content_row += 1
                        
                        # Обновляем current_row до максимального значения
                        max_rows = max(
                            len(ad_combinations),
                            len(sitelinks)
                        )
                        current_row = current_row + max_rows + 1
            
            # Добавляем сетку для всех листов медиаплана
            thin_border = Border(left=Side(style='thin'), 
                              right=Side(style='thin'), 
                              top=Side(style='thin'), 
                              bottom=Side(style='thin'))

            for sheet_name in wb.sheetnames:
                if 'Медиаплан' in sheet_name or 'Интересы' in sheet_name:
                    ws = wb[sheet_name]
                    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                        for cell in row:
                            cell.border = thin_border


            
            # Сохраняем файл
            wb.save(output_path)
            print(f"✅ Excel-файл создан: {output_path}")
            return True
            
        except Exception as e:
            print(f"❌ Ошибка при создании Excel-файла: {e}")
            import traceback
            traceback.print_exc()
            return False

    def process_report(self, report: Dict) -> None:
        """Обработать один отчет"""
        print(f"\n{'='*60}")
        print(f"ОБРАБОТКА ОТЧЕТА #{report['id']}")
        print(f"{'='*60}")
        print(f"ID отчета: {report['id']}")
        print(f"ID договора: {report['id_contracts']}")
        print(f"ID заявки: {report['id_requests']}")
        print(f"Номер договора: {report['number_contract']}")
        print(f"Предмет договора: {report['subject_contract']}")
        
        # Загружаем данные из MinIO
        print(f"\n📥 Загрузка данных из MinIO...")
        campaigns_data = self.load_file_from_minio(report['id'], 'campaigns.json')
        
        if not campaigns_data:
            print("❌ Не удалось загрузить campaigns.json")
            return
        
        keywords_data = self.load_file_from_minio(report['id'], f'keywords_traffic_forecast_{report["id"]}.json')
        if not keywords_data:
            print(f"❌ Не удалось загрузить keywords_traffic_forecast_{report['id']}.json")
            return
        
        # Категоризируем кампании
        print(f"\n🔍 Категоризация кампаний...")
        categories = self.categorize_campaigns(campaigns_data, report['campaign_ids'], keywords_data)
        
        if not categories:
            print("⚠ Нет кампаний для обработки")
            return
        
        # Загружаем все необходимые данные
        ads_data = self.load_file_from_minio(report['id'], f'ads_report_{report["id"]}.json')
        if not ads_data:
            print(f"❌ Не удалось загрузить ads_report_{report['id']}.json")
            return
            
        extensions_data = self.load_file_from_minio(report['id'], f'extensions_{report["id"]}.json')
        if not extensions_data:
            print(f"❌ Не удалось загрузить extensions_{report['id']}.json")
            return
            
        sitelinks_data = self.load_file_from_minio(report['id'], f'sitelinks_{report["id"]}.json')
        if not sitelinks_data:
            print(f"❌ Не удалось загрузить sitelinks_{report['id']}.json")
            return
            
        adgroups_data = self.load_file_from_minio(report['id'], f'adgroups_{report["id"]}.json')
        if not adgroups_data:
            print(f"❌ Не удалось загрузить adgroups_{report['id']}.json")
            return
            
        # Создаем Excel-файл
        print(f"\n📊 Создание Excel-файла...")
        report_id = report.get('id')
        # Генерируем имя файла с датой и временем
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Медиаплан_{timestamp}.xlsx"
        output_path = os.path.join(self.output_folder, filename)

        success = self.create_mediaplan_excel(report['id'], categories, keywords_data, ads_data, extensions_data, sitelinks_data, adgroups_data, output_path)
        
        if success:
            # читаем файл и переводим в байты
            file = io.BytesIO(open(output_path, 'rb').read())

            # # отправляем в S3
            # s3_file_path = os.getenv('S3_REPORT_PATH')
            filename = f"{report['id']}/" + filename
            # s3_file_path = '/'.join((s3_file_path, filename))
            # self.minio_client.put_object(
            #     self.bucket_name, s3_file_path, file, len(file.getvalue()))
            #
            # # записываем адрес (S3) в БД
            # write_s3path_to_bd(report.get('id'), os.getenv('MEDIAPLAN_COL_NAME'), s3_file_path)

            # удаляем файл
            os.remove(output_path)
            return file, filename

            print(f"✅ Медиаплан успешно создан")
        else:
            print(f"❌ Не удалось создать медиаплан")

    def run(self):
        """Основной метод запуска обработки"""
        print("🚀 Запуск генератора медиапланов...")
        
        # Получаем отчеты для обработки
        reports = self.get_pending_reports()
        
        if not reports:
            print("📭 Нет отчетов со статусом 1 для обработки")
            return
        
        print(f"📋 Найдено отчетов для обработки: {len(reports)}")
        
        # Обрабатываем каждый отчет
        for report in reports:
            try:
                self.process_report(report)
            except Exception as e:
                print(f"❌ Ошибка при обработке отчета {report['id']}: {e}")
                import traceback
                traceback.print_exc()
                continue
        
        print(f"\n✅ Обработка завершена. Обработано отчетов: {len(reports)}")


def generate_mediaplan(report_id):
    """Главная функция"""
    try:
        generator = MediaPlanGenerator()
        from utils.postprocessing_report_file import get_report_by_id
        report = get_report_by_id(report_id)

        file, filename = generator.process_report(report)
        return file, filename

        # generator.run()
    except Exception as e:
        print(f"❌ Критическая ошибка: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    print(generate_mediaplan(16))
